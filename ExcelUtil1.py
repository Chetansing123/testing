import openpyxl
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook =openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def ReadData(file,sheetName,ronum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=ronum,column=columnno).value


def WriteData(file,sheetName,ronum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=ronum, column=columnno).value = data
    workbook.save(file)