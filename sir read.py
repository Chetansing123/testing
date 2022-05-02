import openpyxl

# load the workbook
#workbook = openpyxl.load_workbook("C:\\Users\\cheta\\Downloads\\Read.xlsx")
workbook = openpyxl.load_workbook("C:\\Users\\cheta\\Desktop\\read.xlsx")
# load the sheet
sheet = workbook.active

# count row & column
rows=sheet.max_row
cols= sheet.max_column

# print no.of row & column
print(rows)
print(cols)

# print the Excel values using for loop
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value, end= " ")
    print()